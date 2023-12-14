from django.forms import modelform_factory
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template import loader
from personas.forms import LibroForm
from personas.models import Libro
# Create your views here.

def agregar_libro(request):
    pagina = loader.get_template('agregar_libro.html')
    if request.method == 'GET':
        form = LibroForm
    elif request.method == 'POST':
        form = LibroForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('libreria')
    datos = {'form':form}
    return HttpResponse(pagina.render(datos,request))

def modificar_libro(request, id):
    pagina = loader.get_template('modificar_libro.html')
    libro = get_object_or_404(Libro, pk=id)
    if request.method == 'GET':
        form = LibroForm(instance=libro)
    elif request.method == 'POST':
        form = LibroForm(request.POST, instance=libro)
        if form.is_valid():
            form.save()
            return redirect('libreria')
    datos = {'form': form}
    return HttpResponse(pagina.render(datos, request))

def ver_libro(request, id):
    libro = get_object_or_404(Libro, pk=id)
    datos = {'libro':libro}
    pagina = loader.get_template('ver_libro.html')
    return HttpResponse(pagina.render(datos, request))

def eliminar_libro(request, id):
    libro = get_object_or_404(Libro, pk=id)
    if libro:
        libro.delete()
        return redirect('libreria')



# Vista genérica para mostrar resultados
from django.views.generic.base import TemplateView
# Workbook nos permite crear libros en excel
from openpyxl import Workbook
# Nos devuelve un objeto resultado, en este caso un archivo de excel
from django.http.response import HttpResponse

from django.http import HttpResponse
from openpyxl import Workbook
from .models import Libro  # Asegúrate de importar tu modelo de Libro aquí


from django.http import HttpResponse
from openpyxl import Workbook
from .models import Libro  # Asegúrate de importar tu modelo de Libro aquí

def generar_reporte_libros(request):
    # Obtenemos todos los libros de nuestra base de datos
    libros = Libro.objects.order_by('titulo', 'autor', 'Categoria','Presentacion','fecha_publicacion','sinopsis','disponibilidad' )

    # Creamos el libro de trabajo
    wb = Workbook()
    # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
    ws = wb.active

    # Creamos los encabezados en la hoja de trabajo
    ws['B1'] = 'REPORTE DE LA LIBRERIA'
    ws.merge_cells('B1:I1')

    ws['B3'] = 'ID'
    ws['C3'] = 'Título'
    ws['D3'] = 'Autor'
    ws['E3'] = 'Categoría'
    ws['F3'] = 'Presentación'
    ws['G3'] = 'Fecha de Publicación'
    ws['H3'] = 'Sinopsis'
    ws['I3'] = 'Disponibilidad'

    cont = 4
    # Recorremos el conjunto de libros y vamos escribiendo cada uno de los datos en las celdas
    for libro in libros:
        ws.cell(row=cont, column=2).value = libro.id
        ws.cell(row=cont, column=3).value = libro.titulo
        ws.cell(row=cont, column=4).value = libro.autor
        ws.cell(row=cont, column=5).value = libro.Categoria.Categoria
        ws.cell(row=cont, column=6).value = libro.Presentacion.Presentacion
        ws.cell(row=cont, column=7).value = libro.fecha_publicacion
        ws.cell(row=cont, column=8).value = libro.sinopsis
        ws.cell(row=cont, column=9).value = 'Activa' if libro.disponibilidad else 'Inactiva'
        cont = cont+1

    # Establecemos el nombre del archivo
    nombre_archivo = "ReporteLibrosExcel.xlsx"

    # Definimos que el tipo de respuesta a devolver es un archivo de Microsoft Excel
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido

    # Guardamos el libro de trabajo en la respuesta HTTP
    wb.save(response)

    return response

